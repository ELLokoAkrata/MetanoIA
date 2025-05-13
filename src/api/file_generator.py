"""
Módulo para la generación de archivos utilizando herramientas (tools) de Groq.

Este módulo proporciona las definiciones de herramientas y funciones para generar
archivos en diferentes formatos (JSON, Python, Markdown, TXT, CSV, HTML, CSS, JS, Excel)
utilizando la API de Groq.
"""
import os
import json
import tempfile
import re
import csv
import io
from datetime import datetime
from typing import Dict, Any, List, Optional, Union, Tuple

# Definiciones de tipos de archivos soportados
FILE_TYPES = {
    'json': {
        'extension': '.json',
        'mime_type': 'application/json',
        'description': 'Archivo JSON para datos estructurados',
        'content_type': 'object',
        'binary': False
    },
    'python': {
        'extension': '.py',
        'mime_type': 'text/x-python',
        'description': 'Script Python ejecutable',
        'content_type': 'string',
        'binary': False
    },
    'markdown': {
        'extension': '.md',
        'mime_type': 'text/markdown',
        'description': 'Documento Markdown para texto formateado',
        'content_type': 'string',
        'binary': False
    },
    'text': {
        'extension': '.txt',
        'mime_type': 'text/plain',
        'description': 'Archivo de texto plano',
        'content_type': 'string',
        'binary': False
    },
    'csv': {
        'extension': '.csv',
        'mime_type': 'text/csv',
        'description': 'Archivo CSV para datos tabulares',
        'content_type': 'array',  # Array de arrays o array de objetos
        'binary': False
    },
    'html': {
        'extension': '.html',
        'mime_type': 'text/html',
        'description': 'Documento HTML para páginas web',
        'content_type': 'string',
        'binary': False
    },
    'css': {
        'extension': '.css',
        'mime_type': 'text/css',
        'description': 'Hoja de estilos CSS',
        'content_type': 'string',
        'binary': False
    },
    'javascript': {
        'extension': '.js',
        'mime_type': 'text/javascript',
        'description': 'Script JavaScript',
        'content_type': 'string',
        'binary': False
    },
    'excel': {
        'extension': '.xlsx',
        'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'description': 'Hoja de cálculo Excel',
        'content_type': 'array',  # Array de arrays o array de objetos
        'binary': True,
        'requires_library': 'openpyxl'
    }
}

class FileGenerator:
    """
    Clase para generar archivos en diferentes formatos utilizando herramientas de Groq.
    
    Esta clase proporciona las definiciones de herramientas y funciones para generar
    archivos en diferentes formatos (JSON, Python, Markdown, TXT, CSV, HTML, CSS, JS, Excel).
    
    La clase está diseñada para ser extensible y fácil de mantener, con un enfoque
    modular que permite añadir nuevos tipos de archivos con facilidad.
    """
    
    def __init__(self, temp_dir: Optional[str] = None, logger=None):
        """
        Inicializa el generador de archivos.
        
        Args:
            temp_dir (str, optional): Directorio temporal para guardar los archivos generados.
                                     Si no se proporciona, se utilizará el directorio temporal del sistema.
            logger (logging.Logger, optional): Logger para registrar información.
        """
        self.temp_dir = temp_dir or tempfile.gettempdir()
        self.logger = logger
        
        # Crear directorio temporal si no existe
        if not os.path.exists(self.temp_dir):
            os.makedirs(self.temp_dir)
        
        # Verificar disponibilidad de bibliotecas opcionales
        self.available_libraries = self._check_optional_libraries()
            
        if self.logger:
            self.logger.info(f"FileGenerator inicializado con directorio temporal: {self.temp_dir}")
            self.logger.info(f"Bibliotecas opcionales disponibles: {', '.join(self.available_libraries) if self.available_libraries else 'ninguna'}")
    
    def _check_optional_libraries(self) -> List[str]:
        """
        Verifica la disponibilidad de bibliotecas opcionales para tipos de archivos avanzados.
        
        Returns:
            List[str]: Lista de nombres de bibliotecas disponibles.
        """
        available = []
        
        # Verificar openpyxl para archivos Excel
        try:
            import openpyxl
            available.append('openpyxl')
        except ImportError:
            if self.logger:
                self.logger.warning("Biblioteca 'openpyxl' no disponible. La generación de archivos Excel estará deshabilitada.")
        
        return available
    
    def get_tools_definitions(self) -> List[Dict[str, Any]]:
        """
        Obtiene las definiciones de herramientas para la generación de archivos.
        Genera dinámicamente las definiciones basadas en los tipos de archivos soportados.
        
        Returns:
            List[Dict[str, Any]]: Lista de definiciones de herramientas.
        """
        tools = []
        
        # Generar definiciones para cada tipo de archivo soportado
        for file_type, config in FILE_TYPES.items():
            # Omitir tipos que requieren bibliotecas no disponibles
            if config.get('requires_library') and config['requires_library'] not in self.available_libraries:
                continue
                
            # Determinar el tipo de contenido para el parámetro
            content_schema = {
                "description": f"El contenido del archivo {file_type}"
            }
            
            if config['content_type'] == 'object':
                content_schema["type"] = "object"
            elif config['content_type'] == 'array':
                content_schema["type"] = "array"
            else:  # string por defecto
                content_schema["type"] = "string"
            
            # Crear la definición de la herramienta
            tool = {
                "type": "function",
                "function": {
                    "name": f"generate_{file_type}_file",
                    "description": f"Genera un archivo {file_type.upper()} ({config['extension']}) - {config['description']}",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "content": content_schema,
                            "filename": {
                                "type": "string",
                                "description": "Nombre del archivo a generar (sin extensión)"
                            }
                        },
                        "required": ["content", "filename"]
                    }
                }
            }
            
            tools.append(tool)
        
        return tools
    
    def get_available_functions(self) -> Dict[str, callable]:
        """
        Obtiene un diccionario con las funciones disponibles para las herramientas.
        Genera dinámicamente el mapeo de funciones basado en los tipos de archivos soportados.
        
        Returns:
            Dict[str, callable]: Diccionario con las funciones disponibles.
        """
        functions = {}
        
        # Añadir funciones para cada tipo de archivo soportado
        for file_type, config in FILE_TYPES.items():
            # Omitir tipos que requieren bibliotecas no disponibles
            if config.get('requires_library') and config['requires_library'] not in self.available_libraries:
                continue
                
            # Usar el método genérico para todos los tipos
            functions[f"generate_{file_type}_file"] = lambda content, filename, ft=file_type: self.generate_file(content, filename, ft)
        
        return functions
    
    def generate_file(self, content: Any, filename: str, file_type: str) -> Dict[str, Any]:
        """
        Método genérico para generar archivos de cualquier tipo soportado.
        
        Args:
            content (Any): Contenido del archivo (puede ser string, dict, list, etc. según el tipo).
            filename (str): Nombre del archivo sin extensión.
            file_type (str): Tipo de archivo a generar (json, python, markdown, etc.).
            
        Returns:
            Dict[str, Any]: Información del archivo generado.
        """
        # Verificar que el tipo de archivo esté soportado
        if file_type not in FILE_TYPES:
            error_msg = f"Tipo de archivo no soportado: {file_type}"
            if self.logger:
                self.logger.error(error_msg)
            return {"success": False, "error": error_msg}
        
        # Verificar bibliotecas requeridas
        if FILE_TYPES[file_type].get('requires_library') and FILE_TYPES[file_type]['requires_library'] not in self.available_libraries:
            error_msg = f"No se puede generar archivo {file_type}. Biblioteca requerida '{FILE_TYPES[file_type]['requires_library']}' no disponible."
            if self.logger:
                self.logger.error(error_msg)
            return {"success": False, "error": error_msg}
        
        # Asegurar que el nombre del archivo sea seguro
        safe_filename = self._sanitize_filename(filename)
        
        # Añadir extensión si no la tiene
        extension = FILE_TYPES[file_type]['extension']
        if not safe_filename.lower().endswith(extension.lower()):
            safe_filename += extension
        
        # Crear ruta completa
        file_path = os.path.join(self.temp_dir, safe_filename)
        
        try:
            # Procesar y guardar el contenido según el tipo de archivo
            if file_type == 'json':
                self._save_json_file(content, file_path)
            elif file_type == 'csv':
                self._save_csv_file(content, file_path)
            elif file_type == 'excel':
                self._save_excel_file(content, file_path)
            else:
                # Para archivos de texto (python, markdown, text, html, css, js)
                self._save_text_file(content, file_path)
            
            if self.logger:
                self.logger.info(f"Archivo {file_type.upper()} generado: {file_path}")
            
            return {
                "success": True,
                "file_path": file_path,
                "file_type": file_type,
                "file_name": os.path.basename(file_path),
                "mime_type": FILE_TYPES[file_type]['mime_type'],
                "message": f"Archivo {file_type.upper()} '{safe_filename}' generado correctamente."
            }
            
        except Exception as e:
            error_msg = f"Error al generar archivo {file_type}: {str(e)}"
            if self.logger:
                self.logger.error(error_msg)
                self.logger.exception("Detalles del error:")
            
            return {
                "success": False,
                "error": error_msg
            }
    
    def _save_json_file(self, content: Dict[str, Any], file_path: str) -> None:
        """
        Guarda contenido en formato JSON.
        
        Args:
            content (Dict[str, Any]): Contenido a guardar.
            file_path (str): Ruta del archivo.
        """
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(content, f, indent=2, ensure_ascii=False)
    
    def _save_text_file(self, content: str, file_path: str) -> None:
        """
        Guarda contenido como texto plano.
        
        Args:
            content (str): Contenido a guardar.
            file_path (str): Ruta del archivo.
        """
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(content)
    
    def _save_csv_file(self, content: Union[List[List[Any]], List[Dict[str, Any]]], file_path: str) -> None:
        """
        Guarda contenido en formato CSV.
        
        Args:
            content (Union[List[List[Any]], List[Dict[str, Any]]]): Datos para el CSV.
                Puede ser una lista de listas (filas y columnas) o una lista de diccionarios.
            file_path (str): Ruta del archivo.
        """
        with open(file_path, 'w', encoding='utf-8', newline='') as f:
            # Determinar si es una lista de diccionarios o una lista de listas
            if content and isinstance(content[0], dict):
                # Lista de diccionarios
                fieldnames = list(content[0].keys())
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(content)
            else:
                # Lista de listas
                writer = csv.writer(f)
                writer.writerows(content)
    
    def _save_excel_file(self, content: Union[List[List[Any]], Dict[str, List[List[Any]]]], file_path: str) -> None:
        """
        Guarda contenido en formato Excel (XLSX).
        
        Args:
            content: Puede ser una lista de listas (una hoja) o un diccionario de listas de listas (múltiples hojas).
            file_path (str): Ruta del archivo.
        """
        try:
            import openpyxl
            from openpyxl import Workbook
            
            wb = Workbook()
            
            # Si es un diccionario, cada clave es una hoja
            if isinstance(content, dict):
                # Eliminar la hoja por defecto
                default_sheet = wb.active
                wb.remove(default_sheet)
                
                # Crear cada hoja con su contenido
                for sheet_name, sheet_data in content.items():
                    ws = wb.create_sheet(title=sheet_name)
                    for row_idx, row in enumerate(sheet_data, 1):
                        for col_idx, value in enumerate(row, 1):
                            ws.cell(row=row_idx, column=col_idx, value=value)
            else:
                # Una sola hoja con los datos
                ws = wb.active
                for row_idx, row in enumerate(content, 1):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
            
            wb.save(file_path)
            
        except ImportError:
            raise ImportError("La biblioteca 'openpyxl' es necesaria para generar archivos Excel.")
    
    def generate_python_file(self, content: str, filename: str) -> Dict[str, Any]:
        """
        Genera un archivo Python y devuelve la información para su descarga.
        
        Args:
            content (str): Contenido del archivo Python (código).
            filename (str): Nombre del archivo sin extensión.
            
        Returns:
            Dict[str, Any]: Información del archivo generado.
        """
        # Asegurar que el nombre del archivo sea seguro
        safe_filename = self._sanitize_filename(filename)
        
        # Añadir extensión si no la tiene
        if not safe_filename.endswith('.py'):
            safe_filename += '.py'
        
        # Crear ruta completa
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        file_path = os.path.join(self.temp_dir, f"{safe_filename}")
        
        try:
            # Guardar contenido en el archivo
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(content)
            
            if self.logger:
                self.logger.info(f"Archivo Python generado: {file_path}")
            
            return {
                "success": True,
                "file_path": file_path,
                "file_type": "python",
                "file_name": os.path.basename(file_path),
                "message": f"Archivo Python '{safe_filename}' generado correctamente."
            }
        except Exception as e:
            error_msg = f"Error al generar archivo Python: {str(e)}"
            if self.logger:
                self.logger.error(error_msg)
            
            return {
                "success": False,
                "error": error_msg
            }
    
    def generate_markdown_file(self, content: str, filename: str) -> Dict[str, Any]:
        """
        Genera un archivo Markdown y devuelve la información para su descarga.
        
        Args:
            content (str): Contenido del archivo Markdown.
            filename (str): Nombre del archivo sin extensión.
            
        Returns:
            Dict[str, Any]: Información del archivo generado.
        """
        # Asegurar que el nombre del archivo sea seguro
        safe_filename = self._sanitize_filename(filename)
        
        # Añadir extensión si no la tiene
        if not safe_filename.endswith('.md'):
            safe_filename += '.md'
        
        # Crear ruta completa
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        file_path = os.path.join(self.temp_dir, f"{safe_filename}")
        
        try:
            # Guardar contenido en el archivo
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(content)
            
            if self.logger:
                self.logger.info(f"Archivo Markdown generado: {file_path}")
            
            return {
                "success": True,
                "file_path": file_path,
                "file_type": "markdown",
                "file_name": os.path.basename(file_path),
                "message": f"Archivo Markdown '{safe_filename}' generado correctamente."
            }
        except Exception as e:
            error_msg = f"Error al generar archivo Markdown: {str(e)}"
            if self.logger:
                self.logger.error(error_msg)
            
            return {
                "success": False,
                "error": error_msg
            }
    
    def generate_text_file(self, content: str, filename: str) -> Dict[str, Any]:
        """
        Genera un archivo de texto y devuelve la información para su descarga.
        
        Args:
            content (str): Contenido del archivo de texto.
            filename (str): Nombre del archivo sin extensión.
            
        Returns:
            Dict[str, Any]: Información del archivo generado.
        """
        # Asegurar que el nombre del archivo sea seguro
        safe_filename = self._sanitize_filename(filename)
        
        # Añadir extensión si no la tiene
        if not safe_filename.endswith('.txt'):
            safe_filename += '.txt'
        
        # Crear ruta completa
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        file_path = os.path.join(self.temp_dir, f"{safe_filename}")
        
        try:
            # Guardar contenido en el archivo
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(content)
            
            if self.logger:
                self.logger.info(f"Archivo de texto generado: {file_path}")
            
            return {
                "success": True,
                "file_path": file_path,
                "file_type": "text",
                "file_name": os.path.basename(file_path),
                "message": f"Archivo de texto '{safe_filename}' generado correctamente."
            }
        except Exception as e:
            error_msg = f"Error al generar archivo de texto: {str(e)}"
            if self.logger:
                self.logger.error(error_msg)
            
            return {
                "success": False,
                "error": error_msg
            }
    
    def _sanitize_filename(self, filename: str) -> str:
        """
        Sanitiza un nombre de archivo para evitar problemas de seguridad.
        
        Args:
            filename (str): Nombre de archivo a sanitizar.
            
        Returns:
            str: Nombre de archivo sanitizado.
        """
        # Reemplazar caracteres no permitidos con guiones bajos
        safe_name = re.sub(r'[\\/*?:"<>|]', '_', filename)
        # Eliminar espacios al inicio y final
        safe_name = safe_name.strip()
        # Limitar la longitud del nombre
        if len(safe_name) > 100:
            name_part, ext_part = os.path.splitext(safe_name)
            safe_name = name_part[:96] + ext_part if ext_part else name_part[:100]
        # Si está vacío, usar un nombre predeterminado
        if not safe_name:
            safe_name = f"archivo_{datetime.now().strftime('%Y%m%d%H%M%S')}"
        
        return safe_name
        
    def detect_file_type(self, content: Any, filename: str) -> str:
        """
        Detecta automáticamente el tipo de archivo basado en el contenido y nombre.
        
        Args:
            content (Any): Contenido del archivo.
            filename (str): Nombre sugerido para el archivo.
            
        Returns:
            str: Tipo de archivo detectado (json, python, etc.)
        """
        # Verificar extensión en el nombre de archivo
        _, ext = os.path.splitext(filename)
        ext = ext.lower()
        
        # Mapeo de extensiones a tipos de archivo
        ext_to_type = {config['extension']: file_type for file_type, config in FILE_TYPES.items()}
        
        # Si la extensión está en nuestro mapeo, usar ese tipo
        if ext in ext_to_type:
            return ext_to_type[ext]
        
        # Si no hay extensión, intentar detectar por el contenido
        if not ext:
            # Si es un diccionario o lista, probablemente sea JSON
            if isinstance(content, (dict, list)):
                return 'json'
            
            # Si es una cadena, intentar detectar por el contenido
            if isinstance(content, str):
                # Detectar Python
                if content.startswith('import ') or content.startswith('def ') or content.startswith('class '):
                    return 'python'
                # Detectar HTML
                elif content.strip().startswith('<') and ('<html' in content.lower() or '<!doctype html' in content.lower()):
                    return 'html'
                # Detectar CSS
                elif '{' in content and '}' in content and (':' in content or '@media' in content):
                    return 'css'
                # Detectar JavaScript
                elif ('function ' in content or 'const ' in content or 'let ' in content) and ('{' in content and '}' in content):
                    return 'javascript'
                # Detectar Markdown
                elif content.startswith('#') or '**' in content or '```' in content:
                    return 'markdown'
                # Detectar CSV
                elif ',' in content and ('\n' in content or '\r' in content):
                    return 'csv'
            
            # Por defecto, texto plano
            return 'text'
        
        # Si la extensión no está en nuestro mapeo, usar texto por defecto
        return 'text'
